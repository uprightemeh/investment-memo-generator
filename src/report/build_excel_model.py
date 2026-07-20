"""
Builds an Excel version of the same model as src/engine/dcf.py and
ratios.py -- but with LIVE FORMULAS instead of Python-computed numbers, so
you can open it, change an assumption (WACC, growth rate, whatever), and
watch every downstream cell recalculate the way a real analyst's model
would. This is deliberately a separate artifact from the PDF: the PDF is
a static "here's the answer" deliverable, the workbook is the "here's the
machine, go turn the dials yourself" deliverable.

Color convention (standard in real financial models):
  blue text  = hardcoded input you're meant to change
  black text = formula (never edit directly)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(color="0000FF")
BLACK = Font(color="000000")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
BOLD = Font(bold=True)
YELLOW = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_FMT = "$#,##0;($#,##0);-"
PCT_FMT = "0.0%"
MULT_FMT = "0.0x"


def header_row(ws, row, labels, start_col=1):
    for i, label in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + i, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")


def build_assumptions_sheet(wb):
    ws = wb.active
    ws.title = "Assumptions"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50

    ws["A1"] = "DCF Assumptions — Costco Wholesale Corporation (COST)"
    ws["A1"].font = Font(bold=True, size=14, color="1F3A5F")

    rows = [
        ("Base Revenue ($)", 254_453_000_000, "FY2024 actual, Source: 10-K FY2024, EDGAR accession 0000909832-24-000031"),
        ("Year 1 Growth Rate", 0.07, "Analyst estimate — edit this cell to test sensitivity"),
        ("Year 2 Growth Rate", 0.065, "Analyst estimate"),
        ("Year 3 Growth Rate", 0.06, "Analyst estimate"),
        ("Year 4 Growth Rate", 0.055, "Analyst estimate"),
        ("Year 5 Growth Rate", 0.05, "Analyst estimate"),
        ("EBIT Margin", 9_285_000_000 / 254_453_000_000, "FY2024 actual operating income / revenue (exact, matches Python engine)"),
        ("Tax Rate", 0.25, "Approximate blended statutory rate"),
        ("CapEx % of Revenue", 0.018, "Analyst estimate based on historical capex/revenue"),
        ("D&A % of Revenue", 0.012, "Analyst estimate"),
        ("NWC Change % of Revenue Change", 0.01, "Analyst estimate"),
        ("WACC", 0.08, "Key assumption — highly sensitive, see Sensitivity tab"),
        ("Terminal Growth Rate", 0.025, "Key assumption — must be < WACC"),
        ("Net Debt ($)", -10_000_000_000, "Negative = net cash position"),
        ("Diluted Shares Outstanding", 443_000_000, "FY2024 actual"),
        ("Current Share Price ($)", 920.00, "Illustrative — replace with live market price"),
    ]

    start_row = 3
    for i, (label, value, note) in enumerate(rows):
        r = start_row + i
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=value)
        cell.font = BLUE
        cell.fill = YELLOW
        cell.border = BORDER
        if "Rate" in label or "Margin" in label or "%" in label or "WACC" in label:
            cell.number_format = PCT_FMT
        elif "Price" in label:
            cell.number_format = "$#,##0.00"
        elif "Shares" in label:
            cell.number_format = "#,##0"
        else:
            cell.number_format = CURRENCY_FMT
        ws.cell(row=r, column=3, value=note).font = Font(italic=True, color="777777", size=9)

    return {
        "base_revenue": "Assumptions!$B$3",
        "growth": [f"Assumptions!$B${4+i}" for i in range(5)],
        "ebit_margin": "Assumptions!$B$9",
        "tax_rate": "Assumptions!$B$10",
        "capex_pct": "Assumptions!$B$11",
        "da_pct": "Assumptions!$B$12",
        "nwc_pct": "Assumptions!$B$13",
        "wacc": "Assumptions!$B$14",
        "terminal_growth": "Assumptions!$B$15",
        "net_debt": "Assumptions!$B$16",
        "shares": "Assumptions!$B$17",
        "current_price": "Assumptions!$B$18",
    }


def build_dcf_sheet(wb, refs):
    ws = wb.create_sheet("DCF Model")
    for col, width in zip("ABCDEFGH", [26, 14, 14, 14, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = width

    ws["A1"] = "DCF Model — Free Cash Flow Projection"
    ws["A1"].font = Font(bold=True, size=14, color="1F3A5F")

    header_row(ws, 3, ["Line Item", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"])

    labels = ["Revenue", "EBIT", "NOPAT", "CapEx", "D&A", "Change in NWC",
              "Free Cash Flow", "Discount Factor", "Present Value of FCF"]
    for i, label in enumerate(labels):
        ws.cell(row=4 + i, column=1, value=label).font = BOLD if label in ("Revenue", "Free Cash Flow") else Font()

    revenue_row, ebit_row, nopat_row, capex_row, da_row, nwc_row, fcf_row, disc_row, pv_row = range(4, 13)

    for year in range(1, 6):
        col = 1 + year  # column B..F
        col_letter = get_column_letter(col)
        prev_col_letter = get_column_letter(col - 1)
        growth_ref = refs["growth"][year - 1]

        # Revenue: base_revenue * (1+g1) for year 1, else prior year revenue * (1+g)
        if year == 1:
            ws.cell(row=revenue_row, column=col, value=f"={refs['base_revenue']}*(1+{growth_ref})")
        else:
            ws.cell(row=revenue_row, column=col, value=f"={prev_col_letter}{revenue_row}*(1+{growth_ref})")

        ws.cell(row=ebit_row, column=col, value=f"={col_letter}{revenue_row}*{refs['ebit_margin']}")
        ws.cell(row=nopat_row, column=col, value=f"={col_letter}{ebit_row}*(1-{refs['tax_rate']})")
        ws.cell(row=capex_row, column=col, value=f"={col_letter}{revenue_row}*{refs['capex_pct']}")
        ws.cell(row=da_row, column=col, value=f"={col_letter}{revenue_row}*{refs['da_pct']}")

        if year == 1:
            ws.cell(row=nwc_row, column=col, value=f"=({col_letter}{revenue_row}-{refs['base_revenue']})*{refs['nwc_pct']}")
        else:
            ws.cell(row=nwc_row, column=col, value=f"=({col_letter}{revenue_row}-{prev_col_letter}{revenue_row})*{refs['nwc_pct']}")

        ws.cell(row=fcf_row, column=col,
                value=f"={col_letter}{nopat_row}+{col_letter}{da_row}-{col_letter}{capex_row}-{col_letter}{nwc_row}")
        ws.cell(row=disc_row, column=col, value=f"=1/(1+{refs['wacc']})^{year}")
        ws.cell(row=pv_row, column=col, value=f"={col_letter}{fcf_row}*{col_letter}{disc_row}")

        for r in range(revenue_row, pv_row + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = BLACK
            cell.border = BORDER
            cell.number_format = "0.0000" if r == disc_row else CURRENCY_FMT

    # --- Valuation bridge ---------------------------------------------
    bridge_row = pv_row + 2
    ws.cell(row=bridge_row, column=1, value="Valuation Bridge").font = Font(bold=True, size=12, color="1F3A5F")

    sum_pv_row = bridge_row + 1
    tv_row = bridge_row + 2
    pv_tv_row = bridge_row + 3
    ev_row = bridge_row + 4
    eq_row = bridge_row + 5
    price_row = bridge_row + 6

    ws.cell(row=sum_pv_row, column=1, value="Sum of PV of FCF (Years 1-5)")
    ws.cell(row=sum_pv_row, column=2, value=f"=SUM(B{pv_row}:F{pv_row})")

    ws.cell(row=tv_row, column=1, value="Terminal Value (Gordon Growth)")
    ws.cell(row=tv_row, column=2,
            value=f"=F{fcf_row}*(1+{refs['terminal_growth']})/({refs['wacc']}-{refs['terminal_growth']})")

    ws.cell(row=pv_tv_row, column=1, value="PV of Terminal Value")
    ws.cell(row=pv_tv_row, column=2, value=f"=B{tv_row}*F{disc_row}")

    ws.cell(row=ev_row, column=1, value="Enterprise Value")
    ws.cell(row=ev_row, column=2, value=f"=B{sum_pv_row}+B{pv_tv_row}")

    ws.cell(row=eq_row, column=1, value="Equity Value")
    ws.cell(row=eq_row, column=2, value=f"=B{ev_row}-{refs['net_debt']}")

    ws.cell(row=price_row, column=1, value="Implied Share Price")
    ws.cell(row=price_row, column=2, value=f"=B{eq_row}/{refs['shares']}")

    for r in range(sum_pv_row, price_row + 1):
        label_cell = ws.cell(row=r, column=1)
        val_cell = ws.cell(row=r, column=2)
        val_cell.font = BLACK if r != price_row else Font(bold=True)
        val_cell.border = BORDER
        val_cell.number_format = "$#,##0.00" if r == price_row else CURRENCY_FMT
        if r == price_row:
            label_cell.font = BOLD

    upside_row = price_row + 2
    ws.cell(row=upside_row, column=1, value="Upside / (Downside) vs. Current Price").font = BOLD
    ws.cell(row=upside_row, column=2,
            value=f"=(B{price_row}-{refs['current_price']})/{refs['current_price']}")
    ws.cell(row=upside_row, column=2).number_format = PCT_FMT
    ws.cell(row=upside_row, column=2).font = BOLD

    return {"implied_price_cell": f"'DCF Model'!$B${price_row}"}


def build_financials_sheet(wb):
    ws = wb.create_sheet("Financials")
    ws.column_dimensions["A"].width = 30
    for col in "BCD":
        ws.column_dimensions[col].width = 16

    ws["A1"] = "Historical Financials — COST ($, FY basis)"
    ws["A1"].font = Font(bold=True, size=14, color="1F3A5F")
    header_row(ws, 3, ["Line Item", "FY2022", "FY2023", "FY2024"])

    data = [
        ("Revenue", 226_954_000_000, 242_290_000_000, 254_453_000_000),
        ("Cost of Goods Sold", 201_402_000_000, 215_214_000_000, 225_954_000_000),
        ("Operating Income", 7_793_000_000, 8_114_000_000, 9_285_000_000),
        ("Net Income", 5_844_000_000, 6_292_000_000, 7_367_000_000),
        ("Stockholders Equity", 20_642_000_000, 22_931_000_000, 26_490_000_000),
        ("Current Assets", 30_664_000_000, 32_351_000_000, 34_326_000_000),
        ("Current Liabilities", 30_413_000_000, 32_991_000_000, 34_549_000_000),
        ("Operating Cash Flow", 7_385_000_000, 9_741_000_000, 11_269_000_000),
        ("CapEx", 3_891_000_000, 4_078_000_000, 4_710_000_000),
    ]
    for i, (label, fy22, fy23, fy24) in enumerate(data):
        r = 4 + i
        ws.cell(row=r, column=1, value=label)
        for col, val in zip("BCD", (fy22, fy23, fy24)):
            cell = ws[f"{col}{r}"]
            cell.value = val
            cell.font = BLUE
            cell.number_format = CURRENCY_FMT
            cell.border = BORDER
    ws.cell(row=13, column=1, value="Source: SEC EDGAR 10-K filings FY2022-FY2024 (XBRL); see README for accession numbers").font = Font(italic=True, size=9, color="777777")
    return {r[0]: 4 + i for i, r in enumerate(data)}  # label -> row


def build_ratios_sheet(wb, fin_rows):
    ws = wb.create_sheet("Ratios")
    ws.column_dimensions["A"].width = 22
    for col in "BCD":
        ws.column_dimensions[col].width = 14

    ws["A1"] = "Financial Ratios (formula-driven from Financials tab)"
    ws["A1"].font = Font(bold=True, size=14, color="1F3A5F")
    header_row(ws, 3, ["Ratio", "FY2022", "FY2023", "FY2024"])

    F = fin_rows
    ratio_formulas = [
        ("Gross Margin", lambda c: f"=(Financials!{c}{F['Revenue']}-Financials!{c}{F['Cost of Goods Sold']})/Financials!{c}{F['Revenue']}"),
        ("Operating Margin", lambda c: f"=Financials!{c}{F['Operating Income']}/Financials!{c}{F['Revenue']}"),
        ("Net Margin", lambda c: f"=Financials!{c}{F['Net Income']}/Financials!{c}{F['Revenue']}"),
        ("ROE", lambda c: f"=Financials!{c}{F['Net Income']}/Financials!{c}{F['Stockholders Equity']}"),
        ("Current Ratio", lambda c: f"=Financials!{c}{F['Current Assets']}/Financials!{c}{F['Current Liabilities']}"),
        ("FCF Margin", lambda c: f"=(Financials!{c}{F['Operating Cash Flow']}-Financials!{c}{F['CapEx']})/Financials!{c}{F['Revenue']}"),
    ]
    for i, (label, _) in enumerate(ratio_formulas):
        ws.cell(row=4 + i, column=1, value=label)

    for i, (label, formula_fn) in enumerate(ratio_formulas):
        r = 4 + i
        for col in "BCD":
            cell = ws[f"{col}{r}"]
            cell.value = formula_fn(col)
            cell.font = BLACK
            cell.border = BORDER
            cell.number_format = "0.0x" if label == "Current Ratio" else PCT_FMT

    # YoY revenue growth, needs prior column -- only computable for FY2023, FY2024
    growth_row = 4 + len(ratio_formulas)
    ws.cell(row=growth_row, column=1, value="Revenue Growth YoY")
    ws[f"B{growth_row}"] = "n/a (no prior year)"
    ws[f"B{growth_row}"].font = Font(italic=True, color="777777")
    ws[f"C{growth_row}"] = f"=(Financials!C{F['Revenue']}-Financials!B{F['Revenue']})/Financials!B{F['Revenue']}"
    ws[f"D{growth_row}"] = f"=(Financials!D{F['Revenue']}-Financials!C{F['Revenue']})/Financials!C{F['Revenue']}"
    for col in "CD":
        ws[f"{col}{growth_row}"].font = BLACK
        ws[f"{col}{growth_row}"].number_format = PCT_FMT
        ws[f"{col}{growth_row}"].border = BORDER


def build_comps_sheet(wb, dcf_refs):
    ws = wb.create_sheet("Comps")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    for col in "CDEFG":
        ws.column_dimensions[col].width = 13

    ws["A1"] = "Comparable Companies"
    ws["A1"].font = Font(bold=True, size=14, color="1F3A5F")
    header_row(ws, 3, ["Ticker", "Company", "EV/EBITDA", "P/E", "EV/Revenue", "Rev Growth", "Op Margin"])

    comps = [
        ("WMT", "Walmart Inc.", 15.2, 28.4, 1.1, 0.056, 0.044),
        ("TGT", "Target Corp.", 8.1, 16.3, 0.6, 0.021, 0.055),
        ("BJ", "BJ's Wholesale Club", 11.4, 19.8, 0.7, 0.041, 0.033),
        ("DG", "Dollar General Corp.", 9.7, 17.5, 0.9, 0.032, 0.049),
    ]
    for i, row_data in enumerate(comps):
        r = 4 + i
        for j, val in enumerate(row_data):
            cell = ws.cell(row=r, column=1 + j, value=val)
            cell.font = BLUE
            cell.border = BORDER
            if j in (2, 3):
                cell.number_format = MULT_FMT
            elif j == 4:
                cell.number_format = MULT_FMT
            elif j in (5, 6):
                cell.number_format = PCT_FMT

    median_row = 4 + len(comps) + 1
    ws.cell(row=median_row, column=2, value="Peer Median").font = BOLD
    for col_idx, col in zip(range(3, 8), "CDEFG"):
        cell = ws.cell(row=median_row, column=col_idx, value=f"=MEDIAN({col}4:{col}7)")
        cell.font = BLACK
        cell.border = BORDER
        cell.number_format = MULT_FMT if col_idx <= 5 else PCT_FMT

    ws.cell(row=median_row + 2, column=1, value="Implied share price (DCF, for reference):").font = Font(italic=True)
    ws.cell(row=median_row + 2, column=3, value=f"={dcf_refs['implied_price_cell']}")
    ws.cell(row=median_row + 2, column=3).number_format = "$#,##0.00"
    ws.cell(row=median_row + 2, column=3).font = BLACK


def main():
    wb = openpyxl.Workbook()
    refs = build_assumptions_sheet(wb)
    dcf_refs = build_dcf_sheet(wb, refs)
    fin_rows = build_financials_sheet(wb)
    build_ratios_sheet(wb, fin_rows)
    build_comps_sheet(wb, dcf_refs)
    wb.save("output/COST_dcf_model.xlsx")
    print("Saved output/COST_dcf_model.xlsx")


if __name__ == "__main__":
    main()
